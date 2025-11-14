import pandas as pd
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

# カード情報を読み込む
card_df = pd.read_csv('card_name.csv', encoding='shift-jis')

# 新メニュー用のカード選定と結果テキスト作成
MENU_ID = 1068

# 項目1: あの人の今のあなたへの気持ちは？
ITEM1_CARDS = [14, 3, 36, 11]  # ALOHA, TAPA, POHAKU, KA WAI A KA PILILIKO

# 項目2: 二人の関係はこれからどう変化する？
ITEM2_CARDS = [42, 34, 18, 40]  # ANUENUE, PUA, PUEO, KANE

# 項目3: 恋を実らせるためにあなたがすべきこと
ITEM3_CARDS = [24, 23, 4, 6]  # HO'OPONOPONO, LAKA, MANA, HI'IAKA

# 各カードの詳細情報を元に鑑定文を作成
RESULTS = {
    # 項目1の結果
    (1, 14): f'''あの人を表しているのは【{card_df[card_df['カードID']==14].iloc[0]['名称']}】のカード。{card_df[card_df['カードID']==14].iloc[0]['概要']}今、あの人はあなたに対して純粋な好意を抱いています。それはエゴや依存のない、誠実な愛情です。あなたといると心が穏やかになり、自然体でいられると感じています。まだ恋愛感情とはっきり自覚していないかもしれませんが、あなたの存在が心に温かさをもたらしていることは確かです。この真の愛の芽生えを大切に育てていけば、やがて美しい花が咲くでしょう。''',

    (1, 3): f'''【{card_df[card_df['カードID']==3].iloc[0]['名称']}】のカードが出ました。{card_df[card_df['カードID']==3].iloc[0]['概要']}あの人は、あなたとの関係を大切に思っています。ただし、適度な距離感を保ちながら、お互いを尊重したいという気持ちも持っています。依存し合うのではなく、それぞれが自立した上で支え合える関係を理想としているのです。焦らずに、健全な境界線を保ちながら信頼関係を築いていくことで、より深い絆が生まれるでしょう。''',

    (1, 36): f'''【{card_df[card_df['カードID']==36].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==36].iloc[0]['概要']}あの人は、あなたとの人間関係の質を見極めようとしています。表面的な付き合いではなく、本当にお互いが成長できる関係性を求めているのです。あなたに対して好意は持っていますが、この関係が自分にとって、そしてあなたにとってもプラスになるかを慎重に見定めている段階です。誠実な態度で接し続けることで、その答えは自然と明らかになるでしょう。''',

    (1, 11): f'''表しているのは【{card_df[card_df['カードID']==11].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==11].iloc[0]['概要']}あの人は、あなたをありのままに見つめています。飾らず、偽らず、純粋にあなたという人間を理解しようとしているのです。あなたの本質や内面の美しさに気づき始めており、その魅力に惹かれています。自分を偽ることなく、素直な気持ちで接することで、相手の心にもその真実が映し出され、より深い理解と信頼が生まれるでしょう。''',

    # 項目2の結果
    (2, 42): f'''【{card_df[card_df['カードID']==42].iloc[0]['名称']}】のカードが現れました。{card_df[card_df['カードID']==42].iloc[0]['概要']}二人の関係は、これから大きな変容を遂げていきます。友人関係から恋愛関係へ、あるいは知り合いから特別な存在へと、質的な変化が訪れるでしょう。この変化には、喜びもあれば不安もあるかもしれません。しかし、暗い部分も明るい部分も含めて受け入れることで、美しい虹のような関係が築かれていきます。希望を持って、この変容の時を迎えましょう。''',

    (2, 34): f'''【{card_df[card_df['カードID']==34].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==34].iloc[0]['概要']}二人の関係は、着実に前進していきます。花が咲くように、自然な流れで距離が縮まっていくでしょう。共通の趣味や価値観を通じて、お互いの魅力を発見し合う時間が増えていきます。焦らず、花が成長するように一歩一歩を大切にすることで、やがて美しい恋の花が咲き誇るでしょう。''',

    (2, 18): f'''【{card_df[card_df['カードID']==18].iloc[0]['名称']}】のカードが出ました。{card_df[card_df['カードID']==18].iloc[0]['概要']}二人の関係は、自然の流れに導かれて進んでいきます。無理に操作しようとせず、自然な成り行きに身を任せることが大切です。運命の導きが、二人を正しい方向へと向かわせてくれるでしょう。困難があっても、それは二人の絆を深めるための試練です。プエオ（フクロウ）のように、明るい意識を持って前を向いていれば、必ず救済の光が差し込みます。''',

    (2, 40): f'''【{card_df[card_df['カードID']==40].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==40].iloc[0]['概要']}二人の関係には、新しい創造のエネルギーが満ちています。これまでとは違う、全く新しい形の関係性が生まれようとしているのです。もし過去に失敗や別れがあったとしても、それはもう一度始めるための準備期間だったのです。新しい朝日が昇るように、二人の関係も新たなステージへと進んでいくでしょう。''',

    # 項目3の結果
    (3, 24): f'''【{card_df[card_df['カードID']==24].iloc[0]['名称']}】のカードが現れました。{card_df[card_df['カードID']==24].iloc[0]['概要']}恋を実らせるためには、正直なコミュニケーションが何より大切です。自分の気持ちを素直に伝え、相手の言葉にも真摯に耳を傾けましょう。誤解や行き違いがあれば、すぐに解決する姿勢を持つことです。批判せず、ただ自分の正直な気持ちを述べることで、二人の関係は修復され、より強固な絆で結ばれていきます。''',

    (3, 23): f'''【{card_df[card_df['カードID']==23].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==23].iloc[0]['概要']}恋を実らせるには、直感とインスピレーションを大切にすることです。頭で考えすぎず、心の声に耳を傾けましょう。あなたの内なる直感が、適切なタイミングや言葉を教えてくれます。フラを踊るように、自然で優雅な振る舞いを心がけることで、相手の心にあなたの魅力が響き渡るでしょう。''',

    (3, 4): f'''【{card_df[card_df['カードID']==4].iloc[0]['名称']}】のカードが出ました。{card_df[card_df['カードID']==4].iloc[0]['概要']}恋を実らせるためには、あなた自身のマナ（エネルギー）を高めることが重要です。ポジティブな言葉を使い、感謝の気持ちを持ち、自分自身を大切にしましょう。あなたが輝いていれば、そのエネルギーは自然と相手にも伝わります。内側から溢れる生命力が、恋を成就させる原動力となるのです。''',

    (3, 6): f'''【{card_df[card_df['カードID']==6].iloc[0]['名称']}】のカードです。{card_df[card_df['カードID']==6].iloc[0]['概要']}恋を実らせるには、献身的な姿勢が必要です。ただし、自分を犠牲にするのではなく、相手の幸せを心から願う気持ちを持つことです。望む結果をイメージし、その未来に向かって誠実に行動しましょう。危険な旅路でも恋を成就させたヒイアカのように、困難があっても諦めずに進み続ける勇気が、最終的に恋を実らせる鍵となります。''',
}

print('新メニュー（ID: 1068）のカード情報と鑑定文を作成しました。')
print('\n=== 項目1: あの人の今のあなたへの気持ちは？ ===')
for card_id in ITEM1_CARDS:
    card_info = card_df[card_df['カードID'] == card_id].iloc[0]
    print(f'カード{card_id}: {card_info["名称"]} ({card_info["読み"]}) - {card_info["キーワード"]}')

print('\n=== 項目2: 二人の関係はこれからどう変化する？ ===')
for card_id in ITEM2_CARDS:
    card_info = card_df[card_df['カードID'] == card_id].iloc[0]
    print(f'カード{card_id}: {card_info["名称"]} ({card_info["読み"]}) - {card_info["キーワード"]}')

print('\n=== 項目3: 恋を実らせるためにあなたがすべきこと ===')
for card_id in ITEM3_CARDS:
    card_info = card_df[card_df['カードID'] == card_id].iloc[0]
    print(f'カード{card_id}: {card_info["名称"]} ({card_info["読み"]}) - {card_info["キーワード"]}')

# EXCELファイルを更新
print('\n\nEXCELファイルを更新中...')
file_path = 'ホロホロタロット追加メニュー.xlsx'
wb = load_workbook(file_path)
ws_result = wb['結果']

# メニューID 1068の結果行を削除
rows_to_delete = []
for row_idx, row in enumerate(ws_result.iter_rows(min_row=2, max_row=ws_result.max_row), start=2):
    if row[0].value == MENU_ID:
        rows_to_delete.append(row_idx)

# 逆順で削除（インデックスがずれないように）
for row_idx in reversed(rows_to_delete):
    ws_result.delete_rows(row_idx)
    print(f'削除: 行 {row_idx}')

print(f'削除完了: {len(rows_to_delete)}行')

# 新しいデータを追加
last_row = ws_result.max_row
print(f'\n新しいデータを追加開始（開始行: {last_row + 1}）')

result_count = 0
all_cards = [ITEM1_CARDS, ITEM2_CARDS, ITEM3_CARDS]

for item_idx, item_cards in enumerate(all_cards, start=1):
    for card_idx, card_id in enumerate(item_cards, start=1):
        new_row = last_row + result_count + 1
        menu_id_str = str(int(MENU_ID))
        result_id = f"{menu_id_str}{item_idx:02d}{card_idx:02d}"

        ws_result.cell(new_row, 1, MENU_ID)  # メニューID
        ws_result.cell(new_row, 2, item_idx)  # 項目ID
        ws_result.cell(new_row, 3, card_idx)  # 結果ID
        ws_result.cell(new_row, 4, card_id)  # カード番号
        ws_result.cell(new_row, 5, RESULTS[(item_idx, card_id)])  # 結果本文

        card_name = card_df[card_df['カードID'] == card_id].iloc[0]['名称']
        print(f'追加: 項目{item_idx}, 結果{card_idx} - カード{card_id} ({card_name})')
        result_count += 1

# 保存
wb.save(file_path)
print(f'\n✓ {file_path} への更新が完了しました！')
print(f'合計 {result_count} 件の鑑定文を更新しました。')
